import random
import threading
import urllib.request
import time
from bs4 import BeautifulSoup
from datetime import datetime
import xlrd
from openpyxl import load_workbook, Workbook
import re
lock = threading.Lock();
class DbTest:
    connect=''
    cursor='';
    keywords=[];
    httpTimeOut=10;
    ucindex=7;
    all_UserAgents=['Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',#safari 5.1 – Windows
                    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0',#IE9
                    'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1',#Firefox 4.0.1 – Windows
                    'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11',#Opera 11.11 – Windows
                    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)',#360浏览器
                    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)',#傲游（Maxthon）
                    'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)', #世界之窗（The World） 3.x
                    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11' #chrome
                    ];
    existurlbrands =dict();
    allurlsforlevel=set();
    allurls=set();
    resultList=set()
    def __init__(self):
        self.keywords = ['Birkenstock', 'Birkens', 'red bull', 'Birki', 'ndsi', 'timberland', 'toms', 'monsoon',
                  'abercrombie', 'watch', 'sexy lingerie', 'lululemon', 'ghd', 'ghi', 'nintendo', 'DVD',
                  'GOLF', 'hollister ', 'fitch', 'omega', 'rolex', 'celine', 'louis vuitton', 'tiffany',
                  'gucci', 'chanel','Clouds','Disk','cigarette'];
        self.keywords=['dyson']
    def getWebsitesByExcel(self,path):
        workbook = load_workbook(path)
        sheets = workbook.get_sheet_names()
        booksheet = workbook.get_sheet_by_name(sheets[0])
        rows = booksheet.rows
        i=0
        for row in rows:
            i=i+1
            if i==1:
                continue

            line = [col.value for col in row]
            url=line[0]
            if url != None and url != '':
                self.allurls.add(url)
                self.allurlsforlevel.add('1' + '@' + url + '@' + url + '@' + '1');
    def httprequert(this):
        while this.allurlsforlevel.__len__()!=0:
            try:
                splits =this.allurlsforlevel.pop().split('@');
            except Exception as e:
                print('e',e);
                continue;
            level = splits[0]
            url = splits[1];
            superurl = splits[2]
            runnum = splits[3]
            url=this.parseUrl(url);
            print('thread name:',threading.current_thread().getName(),':',url)
            try:

                endtime = datetime.now();
                print('httprequesttime:', endtime);

                this.ucindex = random.randint(0,this.all_UserAgents.__len__()-1)
                headers=this.all_UserAgents[this.ucindex];
                headers = {'User-Agent':headers};
                print('headers:',headers,'url:',url)
                req = urllib.request.Request(url=url, headers=headers)
                page = urllib.request.urlopen(req,timeout=this.httpTimeOut);
                content = page.read().decode(encoding='utf-8', errors='strict');
                this.parseContent(content,level,superurl);
                this.resultList.add(url)
            except  Exception as e:
                url = url.strip("http://");
                if int(runnum) == 1:
                    print(url, 'httprequest--第一次请求发生异常', e)
                    this.allurlsforlevel.add('1' + '@' + url + '@' + url + '@2');
                else:
                    print(url, 'httprequest 请求发生异常', e,this.ucindex)
                    list = [];
                    list.append('无法访问');
                    this.existurlbrands.update({superurl: list});
                    this.resultList.add(url)


    def parseUrl(self,url):
        if url.find('http://')==-1 and url.find('https://')==-1:
            url='http://'+url;

        if url[url.__len__()-1:url.__len__()]!='/':
            url=url+'/';
        return  url;

    def parseContent(self,content,level,superurl):
        if self.existurlbrands.get(superurl) == None:
            existbrands =[];
        else:
            existbrands=self.existurlbrands.get(superurl);

        for keyword in self.keywords:
            p1 = "\s"+keyword+"$|^"+keyword+"\s|\s"+keyword+"\s"
            pattern = re.compile(p1)
            a_list = pattern.findall(content.lower());
            if a_list.__len__()!=0 and keyword not in existbrands:
                print('存在敏感关键字:', keyword);
                existbrands.append(keyword);
        self.existurlbrands.update({superurl: existbrands});

    def saveResultINexcel(self,datest):
        print('进入保存结果数据程序')
        headers=list();
        headers.append('网址');
        headers.append('分析状态');
        headers.append('分析备注');

        alldatas=list();
        for k, v in datest.existurlbrands.items():
            data=list();
            data.append(k);
            if list(v).__len__() > 0:
                if list(v)[0] == '无法访问':
                    analyze_status = '无法访问';
                    analyze_remark = '无法访问';
                else:
                    analyze_status = '分析不通过';
                    str = '';
                    for a in v:
                        if str == '':
                            str = a;
                        else:
                            str = str + ',' + a;
                    analyze_remark = str;
            else:
                analyze_status = '分析通过';
                analyze_remark = '不存关键字';

            data.append(analyze_status);
            data.append(analyze_remark);
            alldatas.append(data);


        writefile = 'F:/tuofu2017/learn/Python/files/excel/results.xlsx';
        #writefile="F:/learn/Python/info/excel/results.xlsx"
        wb = Workbook()
        ws = wb.worksheets[0];
        ws.append(headers);
        for data in alldatas:
            ws.append(list(data));
        wb.save(writefile);
        print('do success')

if __name__ == '__main__':
    begintime = datetime.now();
    print('begintime:',begintime);
    datest=DbTest();
    readpath="F:/tuofu2017/learn/Python/files/excel/网站.xlsx";
    #readpath="F:/learn/Python/info/excel/website.xlsx"
    datest.getWebsitesByExcel(readpath);

    print('website num:',datest.allurls.__len__());

    mainthreads = []
    n=1.1;
    n=3;
    for i in range(0,int(datest.allurls.__len__()/1.1)):
        t = threading.Thread(target=datest.httprequert)  # 无参数的情况
        t.setName("main thread:"+str(i));
        mainthreads.append(t)

    print('main threads num:', mainthreads.__len__());
    for t in mainthreads:
    # t.setDaemon(True)
       t.start();

    newexistkeywords=dict();
    x=1;
    y=0;
    waitsecond=20;
    while x==1:
        print('allurlsforlevel length',datest.allurlsforlevel.__len__())
        print('resultList length', datest.resultList.__len__())
        print('allurls length', datest.allurls.__len__())
        if datest.allurlsforlevel.__len__()==0 and datest.allurls.__len__()-datest.resultList.__len__()<20:
            if y==0:
               print('----------------------进入休息---------------------------:',datetime.now())
               #time.sleep(waitsecond)
               print('----------------------休息完毕---------------------------:',datetime.now())
            y=1;
            print('allurls size:',datest.allurls.__len__());
            try:
                datest.saveResultINexcel(datest);
                x = 2;
                print('end')
            except Exception as e:
                print('保存数据异常---------------------------------',e)
            datest.allurlsforlevel.clear();

    endtime = datetime.now();
    print('thread endtime:', endtime);