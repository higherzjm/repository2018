import urllib

import requests
from openpyxl import load_workbook, Workbook


#读xlsx
def readxlsx():

    workbook = load_workbook('F:/tuofu2017/learn/Python/files/excel/test.xlsx')
    # booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    columns = booksheet.columns
    # 迭代所有的行
    for row in rows:
        line = [col.value for col in row]
        print(line[0])


def writexlsx():
    headers = list();
    headers.append('姓名');
    headers.append('年龄');
    alldatas=list()
    for x in range(1,888888):
        data = list();
        print(x)
        data.append("张三"+str(x))
        data.append(x)
        alldatas.append(data)

    writefile = 'F:/tuofu2017/learn/Python/files/excel/results2.xlsx';
    # writefile="F:/learn/Python/info/results.xls"
    wb = Workbook()
    ws = wb.worksheets[0];
    ws.append(headers);
    for data in alldatas:
        ws.append(list(data));
    wb.save(writefile);

if __name__ == '__main__':
    #readxlsx()
    writexlsx()

