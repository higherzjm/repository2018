# coding='utf-8'
from datetime import datetime
import xlrd
import xlwt
from shutil import copy
from openpyxl import Workbook
def read_excel_xlrd():
    #TC_workbook = xlrd.open_workbook(r'F:\learn\Python\info\NewCreateWorkbook.xls')
    path=r'F:\tuofu2017\learn\Python\files\平台交易信息2017-08-22.xls';
    path = 'F:/tuofu2017/learn/Python/files/网址.xls';
    TC_workbook = xlrd.open_workbook(path)
    all_sheets_list = TC_workbook.sheet_names()
    print("All sheets name in File:", all_sheets_list,all_sheets_list.__len__())
    length=all_sheets_list.__len__();
    allsheets=[];
    allheaders=[];
    alldatas=[];

    for x in range(0,length):
       eachdatas = [];
       allsheets.append(x);
       mysheet=TC_workbook.sheet_by_index(x);
       print(mysheet.name)
       for rownum  in range(0,mysheet.nrows):
           if rownum==0:
              allheaders.append(mysheet.row_values(rownum));
           print(mysheet.row_values(rownum));
           if rownum!=0:
             eachdatas.append(mysheet.row_values(rownum));
       alldatas.append(eachdatas);

    write_to_existed_file(allheaders,alldatas);


def write_to_existed_file(allheaders,alldatas):
    file='F:/tuofu2017/learn/Python/files/writefile4.xls';
    wb = Workbook()
    ws = wb.worksheets[0];
    ws.append(allheaders[0]);
    datas=alldatas[0];
    for data in  datas:
      ws.append(list(data));
    wb.save(file);
    print('do success')

def changetxt2excel(txtFileName):
    new_XlsxFileName = txtFileName[:-3] + 'xlsx'
    wb = Workbook()
    ws = wb.worksheets[0]
    with open(txtFileName) as fp:
        for line in fp:
            line = line.strip().split(',')
            ws.append(line)
    wb.save(new_XlsxFileName)
    fp.close();

def  writeexcel2():
    a_header = ['name', 'age'];
    a_datas = ['zhang', '111'];
    style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on', num_format_str='#,##0.00')
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
    file = 'F:\learn\Python\info\writefile2.xls';
    wb = xlwt.Workbook()
    ws = wb.add_sheet('A Test Sheet')

    ws.write(0, 0, 'time', style0)
    ws.write(0, 1, 'name', style0)
    ws.write(1, 0, datetime.now(), style1)
    ws.write(1, 1,'张三', style0)
    ws.write(2, 0, '11:20')
    ws.write(2, 1, 'lisi')
    ws.write(3, 0, '8:10')
    ws.write(3, 1, 'wangwu')

    ws = wb.add_sheet('B Test Sheet')
    ws.write(0, 0, 1234.56, style0)
    ws.write(1, 0, datetime.now(), style1)
    ws.write(2, 0, 1)
    ws.write(2, 1, 1)
    ws.write(2, 2, 9)
    wb.save(file)

if __name__ == "__main__":
    #write_to_existed_file()
     read_excel_xlrd();
     #changetxt2excel('F:\learn\Python\info\myfile.txt')
     # writeexcel2();